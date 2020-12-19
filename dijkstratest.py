import time
from tkinter import *
import openpyxl
from collections import defaultdict, namedtuple
from heapq import *
from openpyxl import Workbook
from astropy.table import Table

#------------Lists
stations = []
tupleList = []

biglist = []
"""reading into files"""
book = openpyxl.load_workbook('London Underground Data.xlsx')
sheet = book.active
#------------Functions
"""implementation of dijkstra algorithm"""

def dijkstra(stationlist, origin, destination):                                 # stationlist is a list of edge and wight
                                                                                # Define a dictionary of g to store graph structure
    graph = defaultdict(list)
    for a,b,c in stationlist:                                                   #a-from_station, b-to_station, c-time
        graph[a].append((c,b))

    startNode, visited, mins = [(0,origin,())], set(), {origin: 0}

    """Find the shortest path between that node and every other"""
    while startNode:                                                             # startNode represent the node at the start and its information
        (cost,edge,path) = heappop(startNode)                                    #set it as the new "current node"
        if edge not in visited:
            visited.add(edge)
            path = (edge, path)

                                                                                # If the destination node has been marked visited, then stop. The algorithm has finish.
            if edge == destination:
                return (cost, path)

            for c, edges in graph.get(edge, ()):                                # Otherwise, select the unvisited node that is marked with the smallest distance,
                if edges in visited:                                            # mark the current node as visited and push it into the visited set.
                    continue
                prev = mins.get(edges, None)
                next = cost + c
                if prev is None or next < prev:
                    mins[edges] = next
                    heappush(startNode, (next, edges, path))
        listOfStations = list(mins.items())

    return float("inf"), listOfStations[-1]


def flatten(result):                                                            #flattening the result from the nested tuple to a single big list
    for i in result:
        if isinstance(i, (list, tuple, set)):
            yield from flatten(i)
        else:
            yield i

def output():
    """ This function contains major part of script from dijkstratest.py file """
    """appending the nodes using nested tuples and list"""
    for row in sheet.iter_rows(min_row=1, min_col=2, max_row=757, max_col=4):   # 757 last one, 49 bakerloo
        stationsList = []
        stationsTuple = ()

        for cell in row:
            stationsList.append(cell.value)
            stationsTuple = stationsTuple + (cell.value,)
            #print(cell.value, end=' ')

        if stationsList[2] is not None:
            stations.append(stationsList)
            tupleList.append(stationsTuple)

            reversedList = stationsList[:]
            element0 = reversedList[0]
            reversedList[0] = reversedList[1]
            reversedList[1] = element0

            stations.append(reversedList)
            tupleList.append(tuple(reversedList))

    #print("print stations", stations)
    #print("print tuple_list", tuple_list)


    for element in tupleList:
        biglist.append(element)

    if __name__ == "__main__":

        """reading into the excel file"""
        wb = Workbook()
        wb = openpyxl.load_workbook("London Underground Data.xlsx")
        ws = wb.active

        source=from_var.get()                                           #getting the source node from the user
        if source == ' ':
            print("Starting point is empty, TRY AGAIN")
            aa = "Starting point is empty, TRY AGAIN"
            text_area.insert(END, aa)

        else:
            destination = to_var.get()                                  #getting the destination from the user

            if destination == ' ':
                print("destination is empty, TRY AGAIN")
                aa = "destination is empty, TRY AGAIN"
                text_area.insert(END, aa)

        travel_time = time_var.get()
        get_hour = travel_time[:2]
        if get_hour.isdigit:
            hr = int(get_hour)
            # if not during traveling time
            if hr in [1, 2, 3, 4] or hr > 23:
                print("Bus does not travel at that time or the time is invalid")
            else:
                stationlist = biglist
                print()
                print(source, '--->', destination)
                print()
                print("=== Your journey is: ===")
                result = list(dijkstra(stationlist, source, destination))
    #graph = Graph(biglist)
    #print("result", result)



    a = list(flatten(result))
    #print("print a (flatten result)", a)


    result_stations = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=757, max_col=4):
        result_stations_row = []

        for cell in row:
            result_stations_row.append(cell.value)
            #print("result_stations_row",result_stations_row)

        if result_stations_row[1] in a and result_stations_row[2] in a:
            result_stations.append(result_stations_row)


    print("result_stations:", result_stations)
    #print(" print result_tuple_list", result_tuple_list)

    print()
    totalTime = a[0]
    print("Total time of your journey is:",totalTime, "minutes")                    #prints the total time taken from source to destination
    print()
    numberStation = a[1:]
    print("The number of the stations you will travel by:", len(numberStation), "stations")     #prints the number of station it stops by
    print()
    stat_between_order = reversed(numberStation)
    print("Stations between", source, "and", destination, ":")                                  #prints the name of the station from source to destination
    for item in stat_between_order:
        print(item)


    print()
    #print("Summing minutes after each station:")
    c = result_stations
    minutes = []
    for i in c:
        minutes.append(i[3])
    #print(minutes)

    def sum_minutes(l):                                             #calculating the time to show the counter steps in the result
        total = 0
        min=[]
        for val in l:
            total = total + val
            #print(total)
            min.append(total)
        return min
    list_min=sum_minutes(minutes)
    #print(list_min)

    for i in range(len(c)):
            c[i].append(list_min[i])

    data_rows_new = c
    tt = Table(rows=data_rows_new, names=('Line:', 'From Station:', 'To Station:', 'Steps:', 'Total Time:'))
    print(tt)
    text_area.delete(1.0, END)
    text_area.insert(END, tt)

    stationLabel.configure(text='Your journey is from ' + from_var.get() + ' to ' + to_var.get())
    print()
    print("line_name:", c)


#-------------Initialize Tkinter Window
window = Tk()
window.title('TFL Journey Planner')
window.geometry('850x600')
window.attributes("-alpha",0.95) #this line makes window transparent for now is 95% nonetransparent
#window.resizable(False, False)
#root.wm_attributes('-transparentcolor', 'light grey')

#------------GUI components
top_frame = Frame(window)
top_frame.pack(fill="both") #expand=True

label = Label(top_frame, text="Plan A Journey", bg="red", fg="white", padx=10, pady=10, font=("Cambria", 20))
label.pack(fill="x")

travelLabel= Label(top_frame, text="Enter the name of the station ypu want to travel from :", padx=20,pady=5)
travelLabel.pack(fill='x')

from_var = StringVar()
from_station = Entry(top_frame, textvariable=from_var, width=30)
from_station.pack(padx=10, pady=5)

from_station.insert(0, 'From ')

travelLabel= Label(top_frame, text="Enter the name of the station you want to travel to : ", padx=20,pady=5)
travelLabel.pack(fill='x')

to_var = StringVar()
to_station = Entry(top_frame, textvariable=to_var, width=30)
to_station.pack(padx=10, pady=5)
to_station.insert(0, 'To ')

first_frame = Frame(window)
first_frame.pack(fill="both")

travelLabel= Label(first_frame, text="What time do you want to travel (in 24hrs clock) : ", padx=20,pady=5)
travelLabel.pack(fill='x')

time_var = StringVar()
travelTime= Entry(first_frame, textvariable=time_var, width=10)
travelTime.pack(padx=5, pady=5)

Button = Button(first_frame, text='FIND ROUTE', command=output)
Button.pack(padx=10, pady=5)

#-----Second frame

second_frame = LabelFrame(window, text='Journey', fg='white',bg='black')#, padx=150, pady=50)
second_frame.pack(fill="y", expand=True)

stationLabel = Label(second_frame, text=" ", font='Cambria, 10' , bg="black", fg="white")
stationLabel.pack(padx=10)

text_area = Text(second_frame,  bg="black", fg="white", height=20, width=750)
text_area.pack(fill="both", expand=True, padx=100, pady=20)

"""Button1 = Label(second_frame,text='Station!', padx=10, pady=10)
Button1.grid(row=0, column=0, pady=10)
Button2 = Label(second_frame,text='Line!', padx=10, pady=10)
Button2.grid(row=0, column=1, pady=10)
Button3 = Label(second_frame,text='Travel Time!', padx=10, pady=10)
Button3.grid(row=0, column=2, pady=10)
Button4 = Label(second_frame,text='Total time!', padx=10, pady=10)
Button4.grid(row=0, column=3, pady=10)"""

window.mainloop()
